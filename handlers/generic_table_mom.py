# handlers/generic_table_mom.py — 通用「环比添加」：导入两个文件名末四位为 YYMM 的 xlsx，自动判月份后给本期回填占比/环比/环比增量

from __future__ import annotations

import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

# 文件名末 4 位 YYMM；允许可选 .xlsx 后缀，YY=00..99，MM=01..12
_TAIL_YYMM = re.compile(r"(\d{2})(\d{2})(?:\.xlsx)?$", re.IGNORECASE)

# 三类计算列（按 header 第一行命名归类）
_RATIO_HEADERS = {"占比", "全国占比"}
_MOM_GROWTH_HEADERS = {"环比", "环比增速"}
_MOM_DELTA_HEADERS = {"环比增量", "环比变化"}
_COMPUTED_HEADERS = _RATIO_HEADERS | _MOM_GROWTH_HEADERS | _MOM_DELTA_HEADERS

# 缺失/占位写法
_MISSING = "\\"


def _num(v) -> Optional[float]:
    """松散数值解析：None/NaN/空/非数值/占位符 → None；'12.3%' → 12.3；千分位逗号容忍。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    s = str(v).strip()
    if not s or s in {_MISSING, "—", "-", "/"}:
        return None
    s = s.replace(",", "")
    try:
        if s.endswith("%"):
            return float(s[:-1])
        return float(s)
    except (TypeError, ValueError):
        return None


def _fmt_4dp(x: float) -> str:
    """四位小数字符串。"""
    return f"{x:.4f}"


def _fmt_delta(curr: float, prev: float) -> str:
    """环比增量：差值；两边都是整数则保留整数，否则四位小数。"""
    d = curr - prev
    if abs(curr - round(curr)) < 1e-9 and abs(prev - round(prev)) < 1e-9:
        return str(int(round(d)))
    return _fmt_4dp(d)


def _fmt_mom_growth(curr: Optional[float], prev: Optional[float]) -> str:
    """与 power_table_mom._fmt_mom_growth 一致。"""
    if curr is None or prev is None:
        return _MISSING
    if prev == 0 and curr == 0:
        return "0.0000"
    if prev == 0:
        return "—"
    return _fmt_4dp((curr - prev) / prev)


def parse_yymm(path: str) -> Optional[Tuple[int, int]]:
    """从文件名末 4 位解析 (year, month)。返回如 (2026, 1)；不合法则 None。"""
    base = os.path.basename(path)
    stem = os.path.splitext(base)[0]
    m = _TAIL_YYMM.search(stem)
    if not m:
        return None
    yy, mm = int(m.group(1)), int(m.group(2))
    if mm < 1 or mm > 12:
        return None
    return 2000 + yy, mm


def _classify_header(name: str) -> Optional[str]:
    """返回 'ratio' / 'mom' / 'delta' / None。"""
    if not name:
        return None
    n = str(name).strip()
    if n in _RATIO_HEADERS:
        return "ratio"
    if n in _MOM_GROWTH_HEADERS:
        return "mom"
    if n in _MOM_DELTA_HEADERS:
        return "delta"
    return None


def _is_numeric_column(ws, col_idx: int, max_row: int) -> bool:
    """判定某列在数据区是否以数值为主（>=1 个数值、且数值占非空单元格的多数）。"""
    total = 0
    nums = 0
    for r in range(2, min(max_row, 200) + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        total += 1
        if _num(v) is not None:
            nums += 1
    return nums >= 1 and nums * 2 >= total  # 多数为数值


def _find_metric_col(ws, comp_col: int, computed_cols: set, max_row: int) -> Optional[int]:
    """对计算列向左找首个非计算、且 header 非空、数据为数值主的列；跳过最左 key 列（首列）。"""
    headers_row = ws[1]
    for c in range(comp_col - 1, 1, -1):  # 跳过列 1（key 列）
        if c in computed_cols:
            continue
        h = headers_row[c - 1].value
        if h is None or not str(h).strip():
            continue
        if _is_numeric_column(ws, c, max_row):
            return c
    return None


def _row_key(ws, row: int) -> str:
    v = ws.cell(row=row, column=1).value
    if v is None:
        return ""
    return str(v).strip()


def _load_prev_metric_map(
    ws_prev,
    metric_header: str,
) -> Dict[str, float]:
    """在上期 sheet 中按相同 header 找指标列；按 key 列建 key→数值。"""
    out: Dict[str, float] = {}
    if ws_prev is None:
        return out
    headers = [c.value for c in ws_prev[1]]
    target_idx = None
    for i, h in enumerate(headers, start=1):
        if h is not None and str(h).strip() == metric_header:
            target_idx = i
            break
    if target_idx is None:
        return out
    for r in range(2, ws_prev.max_row + 1):
        key = _row_key(ws_prev, r)
        if not key:
            continue
        v = _num(ws_prev.cell(row=r, column=target_idx).value)
        if v is not None:
            out[key] = v
    return out


def _process_sheet(ws_curr, ws_prev, sheet_name: str) -> Tuple[int, List[str]]:
    """处理本期一个 sheet。返回 (写入单元格数, 警告列表)。"""
    warnings: List[str] = []
    written = 0

    headers = [c.value for c in ws_curr[1]]
    if not headers:
        return 0, warnings

    # 收集计算列：col_idx (1-based) → kind
    computed: Dict[int, str] = {}
    for i, h in enumerate(headers, start=1):
        kind = _classify_header(h)
        if kind:
            computed[i] = kind
    if not computed:
        return 0, warnings

    max_row = ws_curr.max_row

    # 为每个计算列推断「对应指标」列
    metric_for: Dict[int, int] = {}
    metric_header_for: Dict[int, str] = {}
    for ci in computed:
        mi = _find_metric_col(ws_curr, ci, set(computed.keys()), max_row)
        if mi is None:
            warnings.append(f"[{sheet_name}] 列 {headers[ci - 1]!r} 未找到对应指标列，已跳过。")
            continue
        metric_for[ci] = mi
        metric_header_for[ci] = str(headers[mi - 1]).strip()

    if not metric_for:
        return 0, warnings

    # 预算每个指标列的合计（用于占比）
    metric_totals: Dict[int, float] = {}
    for ci, mi in metric_for.items():
        if computed[ci] != "ratio":
            continue
        if mi in metric_totals:
            continue
        s = 0.0
        for r in range(2, max_row + 1):
            v = _num(ws_curr.cell(row=r, column=mi).value)
            if v is not None:
                s += v
        metric_totals[mi] = s

    # 上期 sheet 名匹配：完全相同；找不到则 ws_prev=None，环比/增量都写 \
    prev_maps: Dict[int, Dict[str, float]] = {}
    if ws_prev is not None:
        for ci in metric_for:
            if computed[ci] in {"mom", "delta"}:
                hh = metric_header_for[ci]
                if hh not in prev_maps:
                    prev_maps[hh] = _load_prev_metric_map(ws_prev, hh)
    else:
        warnings.append(f"[{sheet_name}] 上期未找到同名 Sheet，环比/环比增量列写为 {_MISSING!r}。")

    # 逐行写入
    for r in range(2, max_row + 1):
        key = _row_key(ws_curr, r)
        if not key:
            continue
        for ci, kind in computed.items():
            if ci not in metric_for:
                continue
            mi = metric_for[ci]
            curr = _num(ws_curr.cell(row=r, column=mi).value)
            if kind == "ratio":
                total = metric_totals.get(mi, 0.0)
                if curr is None or total == 0:
                    ws_curr.cell(row=r, column=ci).value = _MISSING
                else:
                    ws_curr.cell(row=r, column=ci).value = _fmt_4dp(curr / total)
                written += 1
            elif kind == "mom":
                pv_map = prev_maps.get(metric_header_for[ci], {})
                prev_v = pv_map.get(key) if ws_prev is not None else None
                ws_curr.cell(row=r, column=ci).value = _fmt_mom_growth(curr, prev_v)
                written += 1
            elif kind == "delta":
                pv_map = prev_maps.get(metric_header_for[ci], {})
                prev_v = pv_map.get(key) if ws_prev is not None else None
                if curr is None or prev_v is None:
                    ws_curr.cell(row=r, column=ci).value = _MISSING
                else:
                    ws_curr.cell(row=r, column=ci).value = _fmt_delta(curr, prev_v)
                written += 1

    return written, warnings


def _norm_path(s: str) -> str:
    return os.path.abspath(os.path.expandvars(os.path.expanduser((s or "").strip().strip('"').strip("'"))))


def _fill_one_pair(
    curr_path: str,
    prev_path: str,
    curr_ym: Tuple[int, int],
    prev_ym: Tuple[int, int],
) -> Tuple[int, List[str]]:
    """处理一对文件。返回 (写入单元格数, 明细行)。"""
    details: List[str] = []
    details.append(f"本期：{os.path.basename(curr_path)}（{curr_ym[0]}-{curr_ym[1]:02d}）")
    details.append(f"上期：{os.path.basename(prev_path)}（{prev_ym[0]}-{prev_ym[1]:02d}）")

    wb_curr = load_workbook(curr_path)
    total_written = 0
    try:
        wb_prev = load_workbook(prev_path, read_only=False, data_only=True)
        try:
            prev_sheet_map = {s.title: s for s in wb_prev.worksheets}
            for ws in wb_curr.worksheets:
                ws_prev = prev_sheet_map.get(ws.title)
                w, warns = _process_sheet(ws, ws_prev, ws.title)
                total_written += w
                if w > 0:
                    details.append(f"  · Sheet 「{ws.title}」写入 {w} 个单元格。")
                for msg in warns:
                    details.append(f"  ! {msg}")
        finally:
            wb_prev.close()
        wb_curr.save(curr_path)
    finally:
        wb_curr.close()
    return total_written, details


def _determine_period_order(
    pa: str, pb: str,
) -> Tuple[Optional[str], str, str, str, Tuple[int, int], Tuple[int, int]]:
    """解析两条路径的月份并确定本期/上期。
    返回 (error_msg_or_None, curr_path, prev_path, _, curr_ym, prev_ym)。"""
    ya = parse_yymm(pa)
    yb = parse_yymm(pb)
    if ya is None or yb is None:
        msgs = []
        if ya is None:
            msgs.append(f"无法解析月份：{os.path.basename(pa)}")
        if yb is None:
            msgs.append(f"无法解析月份：{os.path.basename(pb)}")
        return "; ".join(msgs), "", "", "", (0, 0), (0, 0)
    if ya == yb:
        return f"两者月份相同（{ya[0]}-{ya[1]:02d}）", "", "", "", (0, 0), (0, 0)
    if ya > yb:
        return None, pa, pb, "", ya, yb
    return None, pb, pa, "", yb, ya


def fill_generic_mom_from_two(
    path_a: str,
    path_b: str,
) -> Tuple[bool, str, List[str], str]:
    """
    导入两个 xlsx，按文件名末 4 位 YYMM 自动判月份，新者作为本期被原地写回。
    返回 (是否成功, 总述, 明细行, 实际写回路径)。
    """
    pa, pb = _norm_path(path_a), _norm_path(path_b)
    if not pa or not pb:
        return False, "请填写两个文件路径。", [], ""
    if not os.path.isfile(pa):
        return False, f"文件不存在：{pa}", [], ""
    if not os.path.isfile(pb):
        return False, f"文件不存在：{pb}", [], ""

    err, curr_path, prev_path, _, curr_ym, prev_ym = _determine_period_order(pa, pb)
    if err:
        return False, err, [], ""

    total, details = _fill_one_pair(curr_path, prev_path, curr_ym, prev_ym)
    return (
        True,
        f"已完成：本期 {os.path.basename(curr_path)} 共写入 {total} 个单元格。",
        details,
        curr_path,
    )


def fill_generic_mom_from_two_folders(
    folder_a: str,
    folder_b: str,
) -> Tuple[bool, str, List[str], List[str]]:
    """
    导入两个文件夹，按文件夹名末 4 位 YYMM 判月份，自动配对同名 xlsx 并逐个处理。
    返回 (是否成功, 总述, 明细行, 实际写回的文件路径列表)。
    """
    da, db = _norm_path(folder_a), _norm_path(folder_b)
    if not da or not db:
        return False, "请填写两个文件夹路径。", [], []
    if not os.path.isdir(da):
        return False, f"文件夹不存在：{da}", [], []
    if not os.path.isdir(db):
        return False, f"文件夹不存在：{db}", [], []

    err, curr_dir, prev_dir, _, curr_ym, prev_ym = _determine_period_order(da, db)
    if err:
        return False, f"文件夹名月份解析失败：{err}", [], []

    curr_files = {f for f in os.listdir(curr_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")}
    prev_files = {f for f in os.listdir(prev_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")}
    common = sorted(curr_files & prev_files)
    if not common:
        return (
            False,
            "两个文件夹中没有同名的 xlsx 文件。",
            [f"本期文件夹 ({len(curr_files)} 个 xlsx)：{curr_dir}",
             f"上期文件夹 ({len(prev_files)} 个 xlsx)：{prev_dir}"],
            [],
        )

    all_details: List[str] = []
    all_details.append(f"本期文件夹：{curr_dir}（{curr_ym[0]}-{curr_ym[1]:02d}）")
    all_details.append(f"上期文件夹：{prev_dir}（{prev_ym[0]}-{prev_ym[1]:02d}）")
    all_details.append(f"匹配到 {len(common)} 个同名文件，跳过 {len(curr_files - prev_files)} 个仅本期有的文件。")
    targets: List[str] = []
    grand_total = 0

    for fname in common:
        cp = os.path.join(curr_dir, fname)
        pp = os.path.join(prev_dir, fname)
        try:
            w, file_details = _fill_one_pair(cp, pp, curr_ym, prev_ym)
            grand_total += w
            all_details.append(f"· {fname}：写入 {w} 个单元格")
            for d in file_details[2:]:  # 跳前两行（本期/上期标注，文件夹模式已在顶部统一写了）
                all_details.append(f"  {d}")
            targets.append(cp)
        except Exception as e:
            all_details.append(f"· {fname}：处理失败 - {str(e)[:300]}")

    return (
        True,
        f"已完成：{len(common)} 个文件共写入 {grand_total} 个单元格。",
        all_details,
        targets,
    )
