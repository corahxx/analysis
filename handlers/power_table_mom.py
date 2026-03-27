# handlers/power_table_mom.py — 功率段 xlsx 按相邻月份填充「环比」（与 standard00_transform._fmt_mom_growth 一致：四位小数）

from __future__ import annotations

import glob
import os
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

_PERIOD_SUFFIX = re.compile(r"_(\d{6})\.xlsx$", re.IGNORECASE)


def _num(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(",", "")
        if s.endswith("%"):
            return float(s[:-1])
        return float(s)
    except (TypeError, ValueError):
        return None


def _fmt_mom_growth(curr: Optional[float], prev: Optional[float]) -> str:
    if curr is None or prev is None:
        return ""
    if prev == 0 and curr == 0:
        return "0.0000"
    if prev == 0:
        return "—"
    return f"{((curr - prev) / prev):.4f}"


def period_from_filename(path: str) -> Optional[Tuple[int, int]]:
    m = _PERIOD_SUFFIX.search(os.path.basename(path))
    if not m:
        return None
    yyyymm = m.group(1)
    y, mo = int(yyyymm[:4]), int(yyyymm[4:])
    if mo < 1 or mo > 12:
        return None
    return y, mo


def load_count_map(path: str) -> Dict[Tuple[str, str], float]:
    """(省份, 功率段) -> 数量"""
    out: Dict[Tuple[str, str], float] = {}
    xl = pd.ExcelFile(path)
    for sn in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sn, engine="openpyxl")
        if df.empty or "功率段" not in df.columns or "数量" not in df.columns:
            continue
        if "省份" in df.columns:
            for _, r in df.iterrows():
                p = str(r["省份"]).strip() if pd.notna(r.get("省份")) else ""
                seg = str(r["功率段"]).strip() if pd.notna(r.get("功率段")) else ""
                if not p or not seg:
                    continue
                n = _num(r.get("数量"))
                if n is not None:
                    out[(p, seg)] = n
        else:
            p = str(sn).strip()
            for _, r in df.iterrows():
                seg = str(r["功率段"]).strip() if pd.notna(r.get("功率段")) else ""
                if not seg:
                    continue
                n = _num(r.get("数量"))
                if n is not None:
                    out[(p, seg)] = n
    return out


def fill_workbook(path: str, prev_counts: Dict[Tuple[str, str], float]) -> int:
    wb = load_workbook(path)
    filled = 0
    try:
        for ws in wb.worksheets:
            header_row = [c.value for c in ws[1]]
            if not header_row:
                continue
            hmap = {}
            for i, c in enumerate(header_row):
                if c is None:
                    continue
                k = str(c).strip()
                if k:
                    hmap[k] = i
            if "环比" not in hmap or "功率段" not in hmap or "数量" not in hmap:
                continue
            si, qi, mi = hmap["功率段"], hmap["数量"], hmap["环比"]
            has_prov = "省份" in hmap
            pi = hmap.get("省份")
            sn = str(ws.title).strip()
            for r in range(2, ws.max_row + 1):
                seg_c = ws.cell(row=r, column=si + 1).value
                if seg_c is None:
                    continue
                seg_s = str(seg_c).strip()
                if not seg_s:
                    continue
                if has_prov and pi is not None:
                    pv = ws.cell(row=r, column=pi + 1).value
                    prov = str(pv).strip() if pv is not None else ""
                else:
                    prov = sn
                if not prov:
                    continue
                curr = _num(ws.cell(row=r, column=qi + 1).value)
                prev_v = prev_counts.get((prov, seg_s))
                ws.cell(row=r, column=mi + 1).value = _fmt_mom_growth(curr, prev_v)
                filled += 1
        wb.save(path)
    finally:
        wb.close()
    return filled


def run_fill_power_mom_on_folder(folder: str) -> Tuple[bool, str, List[str]]:
    """
    对目录内所有 *_YYYYMM.xlsx 按月份排序，从第二期起依次用上一期填充「环比」列（原地保存）。
    返回 (是否成功, 总述, 明细行)。
    """
    details: List[str] = []
    folder = os.path.abspath(os.path.expandvars(os.path.expanduser(folder.strip())))
    if not folder or not os.path.isdir(folder):
        return False, "路径无效或不是文件夹。", details

    paths = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    items: List[Tuple[int, int, str]] = []
    for p in paths:
        per = period_from_filename(p)
        if per:
            y, m = per
            items.append((y, m, p))
    items.sort(key=lambda t: (t[0], t[1]))
    if len(items) < 2:
        return (
            False,
            "目录中至少需要 2 个文件名以 _YYYYMM.xlsx 结尾的 xlsx（如 功率段分布_202512.xlsx）。",
            [f"已扫描: {len(paths)} 个 xlsx，其中含有效月份后缀: {len(items)} 个。"],
        )

    for i in range(1, len(items)):
        _, _, cur_p = items[i]
        _, _, prev_p = items[i - 1]
        try:
            prev_counts = load_count_map(prev_p)
            n = fill_workbook(cur_p, prev_counts)
            details.append(
                f"已写入：{os.path.basename(cur_p)} ← 上期 {os.path.basename(prev_p)}，共 {n} 个环比单元格。"
            )
        except Exception as e:
            return (
                False,
                f"处理失败：{os.path.basename(cur_p)}",
                details + [str(e)[:500]],
            )

    return True, f"已完成 {len(items) - 1} 个月度文件的环比填充。", details
