# handlers/highway_template.py — 高速公路产品：按业务模板 xlsx 输出（双 Sheet，省份+数量）

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

# 与 00 表「省份」Sheet 中列名、模板 Sheet 名一致
HIGHWAY_METRIC_SHEETS: tuple[str, ...] = (
    "高速公路沿线已建设及预留建设充电停车位服务区",
    "高速公路沿线已建设充电停车位总数",
)

# 业务约定模板（本机开发优先从此读取；不存在时用项目 assets 副本）
HIGHWAY_TEMPLATE_DESKTOP = Path(
    r"C:\Users\HONOR\Desktop\充电标准七张表输出\充电系统标准数据产品数据需求表\模板_高速公路_样例.xlsx"
)


def _assets_dir() -> Path:
    return Path(__file__).resolve().parent.parent / "assets"


def resolve_highway_template_path() -> Optional[Path]:
    if HIGHWAY_TEMPLATE_DESKTOP.is_file():
        return HIGHWAY_TEMPLATE_DESKTOP
    p = _assets_dir() / "模板_高速公路_样例.xlsx"
    if p.is_file():
        return p
    return None


def _num(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return float(v)
    try:
        x = float(str(v).strip().replace(",", ""))
        if pd.isna(x):
            return None
        return x
    except (TypeError, ValueError):
        return None


def build_highway_workbook_bytes(province_df: Optional[pd.DataFrame]) -> BytesIO:
    """
    基于模板生成 02 高速公路 xlsx。
    province_df 为 00 表「省份」Sheet；按列名与 Sheet 名一致填入「数量」列。
    无模板时退回单 Sheet 说明。
    """
    path = resolve_highway_template_path()
    if path is None or load_workbook is None:
        buf = BytesIO()
        pd.DataFrame(
            {
                "说明": [
                    "未找到高速公路模板。请将 模板_高速公路_样例.xlsx 置于："
                    + str(HIGHWAY_TEMPLATE_DESKTOP)
                    + "；或放入项目 assets/模板_高速公路_样例.xlsx。"
                ]
            }
        ).to_excel(
            buf, index=False, engine="openpyxl"
        )
        buf.seek(0)
        return buf

    wb = load_workbook(path)
    for sheet_name in HIGHWAY_METRIC_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        if (
            province_df is not None
            and "省份" in province_df.columns
            and sheet_name in province_df.columns
        ):
            for _, r in province_df.iterrows():
                pname = str(r.get("省份", "")).strip()
                if not pname:
                    continue
                v = _num(r.get(sheet_name))
                if v is None:
                    ws.append([pname, None])
                elif abs(v - round(v)) < 1e-9:
                    ws.append([pname, int(round(v))])
                else:
                    ws.append([pname, v])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
